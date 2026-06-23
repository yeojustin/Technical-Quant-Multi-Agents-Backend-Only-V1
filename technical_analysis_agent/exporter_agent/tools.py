import os
import openpyxl
from fpdf import FPDF
from google.adk.tools import ToolContext
from google.genai import types

def get_field(obj, field_name):
    """Safely retrieves a field from a Pydantic model or a dictionary."""
    if obj is None:
        return None
    if hasattr(obj, field_name):
        return getattr(obj, field_name)
    elif isinstance(obj, dict):
        return obj.get(field_name)
    return None

def format_metric(val) -> str:
    """Formats a metric value to 2 decimal places if it's numeric."""
    if val is None or val == "N/A" or val == "":
        return "N/A"
    try:
        val_float = float(val)
        return f"{val_float:.2f}"
    except (ValueError, TypeError):
        return str(val)

def format_financial_val(val) -> str:
    """Formats a financial statement value (in millions) cleanly with sign placed correctly."""
    if val is None or val == "N/A" or val == "":
        return "N/A"
    try:
        val_float = float(val)
        if val_float < 0:
            return f"-${abs(val_float):,.2f}M"
        else:
            return f"${val_float:,.2f}M"
    except (ValueError, TypeError):
        return str(val)

def format_eps_val(val) -> str:
    """Formats basic EPS cleanly with sign placed correctly."""
    if val is None or val == "N/A" or val == "":
        return "N/A"
    try:
        val_float = float(val)
        if val_float < 0:
            return f"-${abs(val_float):.2f}"
        else:
            return f"${val_float:.2f}"
    except (ValueError, TypeError):
        return str(val)

def clean_pdf_text(text: str) -> str:
    """Cleans text to prevent FPDF unicode character errors (stripping emojis, etc.)."""
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)
    replacements = {
        "\u201c": '"', "\u201d": '"', # smart quotes
        "\u2018": "'", "\u2019": "'", # smart apostrophes
        "\u2013": "-", "\u2014": "-", # dashes
        "\u2022": "-", # bullets
        "\u2192": "->", "\u2190": "<-", # arrows
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
        
    cleaned = []
    for char in text:
        val = ord(char)
        # Keep standard printable ASCII characters
        if 32 <= val <= 126 or char in "\n\r\t":
            cleaned.append(char)
    return "".join(cleaned)

class InstitutionalPDF(FPDF):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.alias_nb_pages()
        self.set_margins(10, 25, 10)
        
    def header(self):
        # Top banner accent line (Slate Blue)
        self.set_fill_color(31, 58, 86)
        self.rect(0, 0, 210, 8, 'F')
        
        # Title and header metadata
        self.set_y(12)
        self.set_font("helvetica", "B", 8)
        self.set_text_color(120, 120, 120)
        self.cell(0, 5, "CONFIDENTIAL // INSTITUTIONAL QUANTITATIVE RESEARCH", new_x="LMARGIN", new_y="NEXT", align="R")
        
        # Border line under header
        self.set_draw_color(220, 220, 220)
        self.line(10, 18, 200, 18)
        self.set_y(22)

    def footer(self):
        # Position at 1.5 cm from bottom
        self.set_y(-15)
        self.set_font("helvetica", "I", 7.5)
        self.set_text_color(150, 150, 150)
        
        # Confidentiality note and page numbers
        self.cell(100, 10, "CONFIDENTIAL - FOR INSTITUTIONAL USE ONLY", align="L")
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="R")

    def draw_score_chart(self, x, y, width, height, scores):
        """Draws a professional horizontal bar chart for scores natively via vectors."""
        # Outer border and background
        self.set_fill_color(250, 250, 250)
        self.rect(x, y, width, height, 'F')
        self.set_draw_color(210, 210, 210)
        self.set_line_width(0.3)
        self.rect(x, y, width, height, 'D')
        
        # Title
        self.set_xy(x + 4, y + 3)
        self.set_font("helvetica", "B", 9)
        self.set_text_color(31, 58, 86)
        self.cell(0, 4, "Key Performance Scores", new_x="LMARGIN", new_y="NEXT")
        
        chart_y_start = y + 8
        chart_height = height - 13
        
        # New Symmetric Layout Coordinates
        label_width = 30
        bar_area_width = 22 # width of positive/negative sides (total 44)
        chart_x_center = x + 34 + bar_area_width # center axis of the bars = x + 56
        
        # Central axis (0 point)
        self.set_draw_color(180, 180, 180)
        self.set_line_width(0.5)
        self.line(chart_x_center, chart_y_start, chart_x_center, chart_y_start + chart_height)
        
        # Light grid lines at -10 and +10
        self.set_draw_color(235, 235, 235)
        self.set_line_width(0.3)
        self.line(chart_x_center - bar_area_width, chart_y_start, chart_x_center - bar_area_width, chart_y_start + chart_height)
        self.line(chart_x_center + bar_area_width, chart_y_start, chart_x_center + bar_area_width, chart_y_start + chart_height)
        
        # Grid line labels (drawn at bottom to avoid overlapping with top-right title)
        self.set_font("helvetica", "", 7)
        self.set_text_color(150, 150, 150)
        self.set_xy(chart_x_center - 2, chart_y_start + chart_height + 0.5)
        self.cell(4, 3, "0", align="C")
        self.set_xy(chart_x_center - bar_area_width - 4, chart_y_start + chart_height + 0.5)
        self.cell(8, 3, "-10", align="C")
        self.set_xy(chart_x_center + bar_area_width - 4, chart_y_start + chart_height + 0.5)
        self.cell(8, 3, "+10", align="C")
        
        num_bars = len(scores)
        bar_spacing = chart_height / num_bars
        bar_height = bar_spacing * 0.6
        
        for i, (label, val, max_val) in enumerate(scores):
            label = clean_pdf_text(label)
            bar_y = chart_y_start + i * bar_spacing + (bar_spacing - bar_height) / 2
            
            # Draw category label
            self.set_font("helvetica", "B", 7.5)
            self.set_text_color(70, 70, 70)
            self.set_xy(x + 3, bar_y)
            self.cell(label_width, bar_height, label, align="L")
            
            # Calculate bar width
            val_clamped = max(-max_val, min(max_val, val))
            bar_w = (val_clamped / max_val) * bar_area_width
            
            # Draw bar
            if val_clamped >= 0:
                self.set_fill_color(39, 174, 96) # Bullish Green
                self.rect(chart_x_center, bar_y, bar_w, bar_height, 'F')
            else:
                self.set_fill_color(192, 57, 43) # Bearish Red
                self.rect(chart_x_center + bar_w, bar_y, -bar_w, bar_height, 'F')
            
            # Draw value label in a dedicated column on the right side
            self.set_font("helvetica", "B", 7.5)
            if val_clamped >= 0:
                self.set_text_color(39, 174, 96)
                val_text = f"+{int(val)}"
            else:
                self.set_text_color(192, 57, 43)
                val_text = f"{int(val)}"
            
            self.set_xy(x + 80, bar_y)
            self.cell(9, bar_height, val_text, align="R")

    def draw_fundamental_trends(self, x, y, width, height, historical_data):
        """Draws professional fundamental trend charts (Revenue line chart, Net Income/FCF grouped bar chart) natively via vectors."""
        # 1. Background & border container
        self.set_fill_color(252, 253, 255)
        self.rect(x, y, width, height, 'F')
        self.set_draw_color(210, 210, 210)
        self.set_line_width(0.3)
        self.rect(x, y, width, height, 'D')
        
        # Container Title
        self.set_xy(x + 5, y + 4)
        self.set_font("helvetica", "B", 11)
        self.set_text_color(31, 58, 86)
        self.cell(0, 5, "Historical Financial Trends (Multi-Year)", new_x="LMARGIN", new_y="NEXT")
        
        if not historical_data:
            self.set_xy(x + 10, y + height / 2 - 2)
            self.set_font("helvetica", "I", 9.5)
            self.set_text_color(120, 120, 120)
            self.cell(0, 5, "No historical financial data available for trend analysis.")
            return

        # Sort historical data chronologically
        data = sorted(historical_data, key=lambda d: int(get_field(d, "fiscal_year") or 0))
        N = len(data)
        
        # --- 1. ANNUAL REVENUE TREND (LINE CHART) ---
        rev_x = x + 10
        rev_w = width - 20  # 170
        rev_y = y + 12
        
        # Calculate bounds for Revenue
        revs = [float(get_field(item, "revenue_millions") or 0) for item in data]
        max_rev = max(revs) if revs else 1.0
        min_rev = min(revs) if revs else 0.0
        
        if max_rev == min_rev:
            y_max_rev = max_rev * 1.2 if max_rev > 0 else 1.0
            y_min_rev = max_rev * 0.8 if max_rev > 0 else -1.0
        else:
            y_max_rev = max_rev + (max_rev - min_rev) * 0.2
            y_min_rev = max(0.0, min_rev - (max_rev - min_rev) * 0.2)
            
        y_range_rev = y_max_rev - y_min_rev
        if y_range_rev == 0:
            y_range_rev = 1.0
            
        self.set_xy(rev_x, rev_y)
        self.set_font("helvetica", "B", 9.5)
        self.set_text_color(80, 80, 80)
        self.cell(rev_w, 5, "Annual Revenue Trend", align="L", new_x="LMARGIN", new_y="NEXT")
        
        chart_y_start = rev_y + 8
        chart_h = 75
        
        # Grid lines (5 grid lines)
        self.set_draw_color(235, 235, 235)
        self.set_line_width(0.2)
        for pct in [0.0, 0.25, 0.5, 0.75, 1.0]:
            gy = chart_y_start + chart_h * pct
            self.line(rev_x + 8, gy, rev_x + rev_w - 5, gy)
            
        points_rev = []
        for i, item in enumerate(data):
            year = int(get_field(item, "fiscal_year") or 0)
            val = float(get_field(item, "revenue_millions") or 0)
            
            if N > 1:
                cx = (rev_x + 12) + i * (rev_w - 24) / (N - 1)
            else:
                cx = rev_x + rev_w / 2
                
            cy = chart_y_start + chart_h - ((val - y_min_rev) / y_range_rev) * chart_h
            points_rev.append((cx, cy, val, year))
            
        self.set_draw_color(52, 152, 219)
        self.set_line_width(1.5)
        for i in range(1, len(points_rev)):
            self.line(points_rev[i-1][0], points_rev[i-1][1], points_rev[i][0], points_rev[i][1])
            
        for cx, cy, val, year in points_rev:
            self.set_fill_color(31, 58, 86)
            self.set_draw_color(255, 255, 255)
            self.set_line_width(0.6)
            self.ellipse(cx - 2.0, cy - 2.0, 4.0, 4.0, 'FD')
            
            self.set_font("helvetica", "B", 8)
            self.set_text_color(50, 50, 50)
            self.set_xy(cx - 10, cy - 4.5)
            self.cell(20, 3.5, f"${val:,.0f}M", align="C")
            
            self.set_font("helvetica", "", 8.5)
            self.set_text_color(100, 100, 100)
            self.set_xy(cx - 8, chart_y_start + chart_h + 2)
            self.cell(16, 3.5, str(year), align="C")
            
        # --- 2. NET INCOME & FREE CASH FLOW (GROUPED BAR CHART) ---
        earn_x = x + 10
        earn_w = width - 20  # 170
        earn_y = y + 115
        
        nis = [float(get_field(item, "net_income_millions") or 0) for item in data]
        fcfs = [float(get_field(item, "free_cash_flow_millions") or 0) for item in data]
        all_vals = nis + fcfs
        
        max_earn = max(all_vals) if all_vals else 1.0
        min_earn = min(all_vals) if all_vals else -1.0
        
        y_max_earn = max(0.0, max_earn) * 1.2 if max_earn != 0 else 1.0
        y_min_earn = min(0.0, min_earn) * 1.2 if min_earn != 0 else -1.0
        y_range_earn = y_max_earn - y_min_earn
        if y_range_earn == 0:
            y_range_earn = 1.0
            
        self.set_xy(earn_x, earn_y)
        self.set_font("helvetica", "B", 9.5)
        self.set_text_color(80, 80, 80)
        self.cell(earn_w - 60, 5, "Net Income & Free Cash Flow Trend", align="L")
        
        # Legend (aligned right, preventing overlap)
        self.set_font("helvetica", "", 7.5)
        self.set_text_color(120, 120, 120)
        self.set_fill_color(39, 174, 96)
        self.rect(earn_x + earn_w - 45, earn_y + 1.2, 4, 3, 'F')
        self.set_xy(earn_x + earn_w - 40, earn_y)
        self.cell(18, 4.5, "Net Income", align="L")
        
        self.set_fill_color(230, 126, 34)
        self.rect(earn_x + earn_w - 20, earn_y + 1.2, 4, 3, 'F')
        self.set_xy(earn_x + earn_w - 15, earn_y)
        self.cell(12, 4.5, "FCF", align="L")
        
        chart_y_start_earn = earn_y + 8
        chart_h_earn = 75
        
        zero_y = chart_y_start_earn + chart_h_earn - ((0.0 - y_min_earn) / y_range_earn) * chart_h_earn
        
        self.set_draw_color(180, 180, 180)
        self.set_line_width(0.5)
        self.line(earn_x + 8, zero_y, earn_x + earn_w - 5, zero_y)
        
        self.set_draw_color(245, 245, 245)
        self.set_line_width(0.2)
        for pct in [0.0, 0.25, 0.5, 0.75, 1.0]:
            gy = chart_y_start_earn + chart_h_earn * pct
            if abs(gy - zero_y) > 2:
                self.line(earn_x + 8, gy, earn_x + earn_w - 5, gy)
        
        bar_w = 6.0
        for i, item in enumerate(data):
            year = int(get_field(item, "fiscal_year") or 0)
            ni_val = float(get_field(item, "net_income_millions") or 0)
            fcf_val = float(get_field(item, "free_cash_flow_millions") or 0)
            
            if N > 1:
                group_cx = (earn_x + 12) + i * (earn_w - 24) / (N - 1)
            else:
                group_cx = earn_x + earn_w / 2
                
            ni_x = group_cx - bar_w - 1.0
            ni_h = (abs(ni_val) / y_range_earn) * chart_h_earn
            if ni_val >= 0:
                ni_y = zero_y - ni_h
                self.set_fill_color(39, 174, 96)
            else:
                ni_y = zero_y
                self.set_fill_color(192, 57, 43)
            self.rect(ni_x, ni_y, bar_w, max(0.2, ni_h), 'F')
            
            fcf_x = group_cx + 1.0
            fcf_h = (abs(fcf_val) / y_range_earn) * chart_h_earn
            if fcf_val >= 0:
                fcf_y = zero_y - fcf_h
                self.set_fill_color(230, 126, 34)
            else:
                fcf_y = zero_y
                self.set_fill_color(211, 84, 0)
            self.rect(fcf_x, fcf_y, bar_w, max(0.2, fcf_h), 'F')
            
            self.set_font("helvetica", "", 8.5)
            self.set_text_color(100, 100, 100)
            self.set_xy(group_cx - 8, chart_y_start_earn + chart_h_earn + 2)
            self.cell(16, 3.5, str(year), align="C")
            
            self.set_font("helvetica", "B", 7.5)
            self.set_text_color(80, 80, 80)
            if ni_val >= 0:
                self.set_xy(ni_x - 5, ni_y - 3.5)
            else:
                self.set_xy(ni_x - 5, ni_y + ni_h + 0.5)
            self.cell(bar_w + 10, 3, f"${ni_val:,.0f}", align="C")
            
            if fcf_val >= 0:
                self.set_xy(fcf_x - 5, fcf_y - 3.5)
            else:
                self.set_xy(fcf_x - 5, fcf_y + fcf_h + 0.5)
            self.cell(bar_w + 10, 3, f"${fcf_val:,.0f}", align="C")


def draw_callout_box(pdf, title, text, r=244, g=247, b=250, border_r=31, border_g=58, border_b=86):
    """Draws a callout box with a thick left border line."""
    title = clean_pdf_text(title)
    text = clean_pdf_text(text)
    pdf.set_fill_color(r, g, b)
    pdf.set_font("helvetica", "", 9)
    width = 190
    
    # Measure heights
    lines = pdf.multi_cell(width - 8, 4.2, text, dry_run=True, output="LINES")
    lines_count = len(lines) if lines else 1
    
    title_height = 5 if title else 0
    box_height = lines_count * 4.2 + title_height + 6
    
    x, y = pdf.get_x(), pdf.get_y()
    
    # Auto page break check
    if y + box_height > pdf.page_break_trigger:
        pdf.add_page()
        x, y = pdf.get_x(), pdf.get_y()
        
    pdf.rect(x, y, width, box_height, 'F')
    
    # Draw left border line
    pdf.set_fill_color(border_r, border_g, border_b)
    pdf.rect(x, y, 2.0, box_height, 'F')
    
    # Render Text
    pdf.set_xy(x + 5, y + 3)
    if title:
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(border_r, border_g, border_b)
        pdf.cell(0, 4, title, new_x="LMARGIN", new_y="NEXT")
        pdf.set_x(x + 5)
        
    pdf.set_font("helvetica", "", 8.5)
    pdf.set_text_color(60, 60, 60)
    pdf.multi_cell(width - 8, 4.2, text, border=0)
    pdf.set_xy(x, y + box_height + 4)

def draw_table_row(pdf, cols, widths, aligns, row_bg_color=None, text_color=(60, 60, 60), is_header=False):
    """Draws a table row with cells wrapped and aligned side-by-side."""
    cols = [clean_pdf_text(col) for col in cols]
    x_start = pdf.get_x()
    y_start = pdf.get_y()
    
    # Calculate cell wrap heights
    max_h = 6.0
    for i, col in enumerate(cols):
        pdf.set_font("helvetica", "B" if is_header else "", 8.5)
        lines = pdf.multi_cell(widths[i], 4, str(col), dry_run=True, output="LINES")
        lines_count = len(lines) if lines else 1
        h_needed = lines_count * 4.2 + 4.0
        if h_needed > max_h:
            max_h = h_needed
            
    # Page break check
    if y_start + max_h > pdf.page_break_trigger:
        pdf.add_page()
        y_start = pdf.get_y()
        x_start = pdf.get_x()
        
    for i, col in enumerate(cols):
        cell_x = x_start + sum(widths[:i])
        pdf.set_xy(cell_x, y_start)
        
        if is_header:
            pdf.set_fill_color(31, 58, 86) # Dark Slate Blue
            pdf.set_draw_color(220, 220, 220)
            pdf.rect(cell_x, y_start, widths[i], max_h, 'FD')
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("helvetica", "B", 8.5)
        else:
            if row_bg_color:
                pdf.set_fill_color(*row_bg_color)
                pdf.set_draw_color(220, 220, 220)
                pdf.rect(cell_x, y_start, widths[i], max_h, 'FD')
            else:
                pdf.set_draw_color(220, 220, 220)
                pdf.rect(cell_x, y_start, widths[i], max_h, 'D')
            pdf.set_text_color(*text_color)
            pdf.set_font("helvetica", "", 8.5)
            
        lines = pdf.multi_cell(widths[i], 4, str(col), dry_run=True, output="LINES")
        lines_count = len(lines) if lines else 1
        padding_top = (max_h - (lines_count * 4.2)) / 2
        
        pdf.set_xy(cell_x, y_start + padding_top)
        pdf.multi_cell(widths[i], 4.2, str(col), align=aligns[i], border=0)
        
    pdf.set_xy(x_start, y_start + max_h)

async def export_report_to_pdf(ticker: str, tool_context: ToolContext) -> str:
    """
    Exports the generated institutional quant report to a PDF file, converting markdown formatting properly.
    
    Args:
        ticker (str): The stock ticker symbol.
    """
    try:
        state = tool_context.session.state
        price_results = state.get("price_feed_results")
        indicator_data = state.get("indicator_final_strategy")
        fundamental_data = state.get("fundamental_results")
        sentiment_data = state.get("sentiment_results")
        thesis = state.get("master_thesis_json")
        
        ticker = clean_pdf_text(ticker or get_field(thesis, "ticker") or "UNKNOWN")
        
        pdf = InstitutionalPDF()
        pdf.add_page()
        
        # --- PAGE 1: TITLE & EXECUTIVE SUMMARY ---
        pdf.set_font("helvetica", "B", 18)
        pdf.set_text_color(31, 58, 86)
        pdf.cell(0, 10, f"Institutional Quant Report: {ticker}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)
        
        # Verdict Status Banner
        verdict = clean_pdf_text(get_field(thesis, "final_verdict") or "N/A")
        conviction = clean_pdf_text(get_field(thesis, "conviction_level") or "N/A")
        
        if "BUY" in verdict:
            bg_color = (39, 174, 96) # Green
        elif "SELL" in verdict:
            bg_color = (192, 57, 43) # Red
        else:
            bg_color = (127, 140, 141) # Grey
            
        banner_y = pdf.get_y()
        pdf.set_fill_color(*bg_color)
        pdf.rect(10, banner_y, 190, 8, 'F')
        pdf.set_font("helvetica", "B", 10)
        pdf.set_text_color(255, 255, 255)
        pdf.set_xy(14, banner_y + 1.5)
        pdf.cell(0, 5, f"Overall Verdict: {verdict}   |   Conviction Level: {conviction}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_y(banner_y + 8 + 4)
        
        # Two-Column Dashboard (Price Feed | Scorecard Chart)
        y_dash = pdf.get_y()
        
        # Left Block: Price Info
        pdf.set_xy(10, y_dash)
        pdf.set_fill_color(250, 250, 250)
        pdf.rect(10, y_dash, 92, 38, 'F')
        pdf.set_draw_color(210, 210, 210)
        pdf.rect(10, y_dash, 92, 38, 'D')
        
        pdf.set_xy(14, y_dash + 3)
        pdf.set_font("helvetica", "B", 9.5)
        pdf.set_text_color(31, 58, 86)
        pdf.cell(0, 5, "Live Market Price Feed", new_x="LMARGIN", new_y="NEXT")
        
        pdf.set_font("helvetica", "", 8.5)
        pdf.set_text_color(70, 70, 70)
        
        curr_price = format_metric(get_field(price_results, "current_price"))
        prev_close = format_metric(get_field(price_results, "previous_close"))
        daily_range = clean_pdf_text(get_field(price_results, "daily_range") or "N/A")
        fifty_two_range = clean_pdf_text(get_field(price_results, "fifty_two_week_range") or "N/A")
        volume = get_field(price_results, "volume") or "N/A"
        
        if isinstance(volume, (int, float)):
            volume_str = f"{volume:,}"
        else:
            volume_str = str(volume)
        volume_str = clean_pdf_text(volume_str)
            
        pdf.set_x(14)
        pdf.cell(42, 4.5, f"Current Price: ${curr_price}")
        pdf.cell(0, 4.5, f"Previous Close: ${prev_close}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_x(14)
        pdf.cell(0, 4.5, f"Daily Range: {daily_range}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_x(14)
        pdf.cell(0, 4.5, f"52-Week Range: {fifty_two_range}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_x(14)
        pdf.cell(0, 4.5, f"Daily Volume: {volume_str}", new_x="LMARGIN", new_y="NEXT")
        
        # Right Block: Horizontal Bar Chart of Scores
        tech_score = get_field(indicator_data, "total_score")
        tech_score_val = int(tech_score / 2) if tech_score is not None else 0
        fund_score_val = get_field(fundamental_data, "fundamental_score") or 0
        sent_score_val = get_field(sentiment_data, "sentiment_score") or 0
        
        scores_list = [
            ("Technical Consensus", tech_score_val, 10),
            ("Fundamental Score", fund_score_val, 10),
            ("Sentiment Score", sent_score_val, 10)
        ]
        
        pdf.draw_score_chart(108, y_dash, 92, 38, scores_list)
        
        # Move cursor under dashboard
        pdf.set_xy(10, y_dash + 43)
        
        # Strategist Thesis Callout
        synthesis = clean_pdf_text(get_field(thesis, "synthesis_summary") or "N/A")
        draw_callout_box(pdf, "Strategist Thesis Synthesis", synthesis)
        
        # --- PAGE 2: FUNDAMENTALS ---
        pdf.add_page()
        pdf.set_font("helvetica", "B", 13)
        pdf.set_text_color(31, 58, 86)
        pdf.cell(0, 8, "Fundamental Health & Valuation", new_x="LMARGIN", new_y="NEXT")
        
        fund_score = clean_pdf_text(get_field(fundamental_data, "fundamental_score") or "N/A")
        valuation_verdict = clean_pdf_text(get_field(fundamental_data, "valuation_verdict") or "N/A")
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(90, 90, 90)
        pdf.cell(0, 5, f"Health Score: {fund_score} (Scale: -10 to +10)  |  Valuation Verdict: {valuation_verdict}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)
        
        catalyst = clean_pdf_text(get_field(fundamental_data, "catalyst_summary") or "N/A")
        draw_callout_box(pdf, "Catalyst Summary", catalyst, r=245, g=247, b=250)
        pdf.ln(2)

        # Detailed Fundamental Score Breakdown Table
        pdf.set_font("helvetica", "B", 10.5)
        pdf.set_text_color(31, 58, 86)
        pdf.cell(0, 6, "Programmatic Fundamental Scoring Breakdown", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1.5)
        
        breakdown_widths = [140, 50]
        breakdown_aligns = ["L", "C"]
        draw_table_row(pdf, ["Fundamental Scoring Component", "Score Contribution"], breakdown_widths, breakdown_aligns, is_header=True)
        
        breakdown = get_field(fundamental_data, "score_breakdown") or {}
        cagr_pts = str(get_field(breakdown, "cagr_points") or 0)
        eps_pts = str(get_field(breakdown, "eps_points") or 0)
        fcf_pts = str(get_field(breakdown, "fcf_points") or 0)
        pe_g_pts = str(get_field(breakdown, "pe_growth_points") or 0)
        peg_pts = str(get_field(breakdown, "peg_points") or 0)
        high_pe_pen = str(get_field(breakdown, "high_pe_penalty") or 0)
        cash_pen = str(get_field(breakdown, "cash_burn_penalty") or 0)
        
        draw_table_row(pdf, ["Revenue CAGR growth points", cagr_pts], breakdown_widths, breakdown_aligns)
        draw_table_row(pdf, ["EPS trend growth points", eps_pts], breakdown_widths, breakdown_aligns)
        draw_table_row(pdf, ["Free Cash Flow health points", fcf_pts], breakdown_widths, breakdown_aligns)
        draw_table_row(pdf, ["Valuation PE growth points", pe_g_pts], breakdown_widths, breakdown_aligns)
        draw_table_row(pdf, ["Valuation PEG points", peg_pts], breakdown_widths, breakdown_aligns)
        draw_table_row(pdf, ["High PE / Low Growth penalty", high_pe_pen], breakdown_widths, breakdown_aligns)
        draw_table_row(pdf, ["Unprofitable cash burn penalty", cash_pen], breakdown_widths, breakdown_aligns)
        pdf.ln(4)
        
        # Valuation Table
        pdf.set_font("helvetica", "B", 10.5)
        pdf.set_text_color(31, 58, 86)
        pdf.cell(0, 6, "Valuation Metrics", new_x="LMARGIN", new_y="NEXT")
        
        metrics = get_field(fundamental_data, "valuation_metrics") or {}
        trailing_pe = format_metric(get_field(metrics, "trailing_pe"))
        forward_pe = format_metric(get_field(metrics, "forward_pe"))
        peg_ratio = format_metric(get_field(metrics, "peg_ratio"))
        price_to_book = format_metric(get_field(metrics, "price_to_book"))
        
        val_widths = [45, 45, 45, 55]
        val_aligns = ["C", "C", "C", "C"]
        draw_table_row(pdf, ["Trailing P/E", "Forward P/E", "PEG Ratio", "Price-to-Book"], val_widths, val_aligns, is_header=True)
        draw_table_row(pdf, [trailing_pe, forward_pe, peg_ratio, price_to_book], val_widths, val_aligns)

        
        pdf.ln(5)
        
        # Historical Statement Table
        pdf.set_font("helvetica", "B", 10.5)
        pdf.set_text_color(31, 58, 86)
        pdf.cell(0, 6, "Multi-Year Financial Statement Summary", new_x="LMARGIN", new_y="NEXT")
        
        cagr = format_metric(get_field(fundamental_data, "revenue_cagr"))
        eps_trend = clean_pdf_text(get_field(fundamental_data, "eps_trend_verdict") or "N/A")
        pdf.set_font("helvetica", "I", 8)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(0, 4.5, f"Revenue CAGR: {cagr}%  |  EPS Trend: {eps_trend}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1.5)
        
        hist_widths = [30, 40, 40, 45, 35]
        hist_aligns = ["C", "R", "R", "R", "R"]
        draw_table_row(pdf, ["Fiscal Year", "Revenue (M)", "Net Income (M)", "Free Cash Flow (M)", "Basic EPS"], hist_widths, hist_aligns, is_header=True)
        
        timeframe_breakdown = get_field(fundamental_data, "timeframe_breakdown") or []
        for i, item in enumerate(timeframe_breakdown):
            fiscal_year = clean_pdf_text(get_field(item, "fiscal_year") or "N/A")
            rev = format_financial_val(get_field(item, "revenue_millions"))
            ni = format_financial_val(get_field(item, "net_income_millions"))
            fcf = format_financial_val(get_field(item, "free_cash_flow_millions"))
            eps = format_eps_val(get_field(item, "eps"))
            
            row_color = (250, 252, 254) if i % 2 == 0 else (255, 255, 255)
            draw_table_row(pdf, [
                fiscal_year,
                rev,
                ni,
                fcf,
                eps
            ], hist_widths, hist_aligns, row_bg_color=row_color)
            
        # Draw native vector trend charts on a dedicated page
        pdf.add_page()
        pdf.draw_fundamental_trends(10, 25, 190, 210, timeframe_breakdown)
            
        # --- PAGE 4: TECHNICALS ---
        pdf.add_page()
        pdf.set_font("helvetica", "B", 13)
        pdf.set_text_color(31, 58, 86)
        pdf.cell(0, 8, "Technical Analysis Momentum & Breakdown", new_x="LMARGIN", new_y="NEXT")
        
        st_rec = clean_pdf_text(get_field(indicator_data, "short_term_recommendation") or "N/A")
        lt_rec = clean_pdf_text(get_field(indicator_data, "long_term_recommendation") or "N/A")
        total_score_val = clean_pdf_text(get_field(indicator_data, "total_score") or 0)
        
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(90, 90, 90)
        pdf.cell(0, 5, f"Short-Term Signal: {st_rec}  |  Long-Term Signal: {lt_rec}  |  Total Score: {total_score_val}/20", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        # Macro Context callout box
        macro_sum = clean_pdf_text(get_field(indicator_data, "macro_summary") or "N/A")
        draw_callout_box(pdf, "Macro Context & Market Structure Synthesis", macro_sum, r=244, g=247, b=250)
        pdf.ln(1)
        
        # Technical narratives
        alligator = clean_pdf_text(get_field(indicator_data, "alligator_status") or "N/A")
        divergence = clean_pdf_text(get_field(indicator_data, "divergence_status") or "N/A")
        
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(31, 58, 86)
        pdf.cell(0, 4.5, "Williams Alligator Analysis:", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("helvetica", "", 8.5)
        pdf.set_text_color(60, 60, 60)
        pdf.multi_cell(0, 4.2, alligator)
        pdf.ln(1.5)
        
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(31, 58, 86)
        pdf.cell(0, 4.5, "RSI Divergence Analysis:", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("helvetica", "", 8.5)
        pdf.set_text_color(60, 60, 60)
        pdf.multi_cell(0, 4.2, divergence)
        pdf.ln(3)

        # Enriched Technical Sub-Indicator Score Breakdown Table
        pdf.set_font("helvetica", "B", 10.5)
        pdf.set_text_color(31, 58, 86)
        pdf.cell(0, 6, "Programmatic Sub-Indicator Scoring Breakdown", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1.5)
        
        tech_breakdown_widths = [140, 50]
        tech_breakdown_aligns = ["L", "C"]
        draw_table_row(pdf, ["Technical Indicator Component", "Score Contribution"], tech_breakdown_widths, tech_breakdown_aligns, is_header=True)
        
        calc_scores = state.get("indicator_calculated_scores") or {}
        breakdown_st = calc_scores.get("breakdown", {}).get("short_term", {})
        breakdown_lt = calc_scores.get("breakdown", {}).get("long_term", {})
        
        for k, v in breakdown_st.items():
            name = f"Short-Term {k.replace('_', ' ').title()}"
            draw_table_row(pdf, [name, str(v)], tech_breakdown_widths, tech_breakdown_aligns)
        for k, v in breakdown_lt.items():
            name = f"Long-Term {k.replace('_', ' ').title()}"
            draw_table_row(pdf, [name, str(v)], tech_breakdown_widths, tech_breakdown_aligns)
        pdf.ln(4)

        
        # Technical Timeframe Table (Enriched with Score Contribution)
        tech_widths = [25, 25, 30, 110]
        tech_aligns = ["C", "C", "C", "L"]
        draw_table_row(pdf, ["Timeframe", "Signal", "Score Contrib", "Technical Justification"], tech_widths, tech_aligns, is_header=True)
        
        tech_breakdown = get_field(indicator_data, "timeframe_breakdown") or []
        for i, item in enumerate(tech_breakdown):
            tf = clean_pdf_text(get_field(item, "timeframe") or "N/A")
            sig = clean_pdf_text(get_field(item, "signal") or "N/A")
            contrib = str(get_field(item, "score_contribution") or 0)
            just = clean_pdf_text(get_field(item, "justification") or "N/A")
            
            row_color = (250, 252, 254) if i % 2 == 0 else (255, 255, 255)
            draw_table_row(pdf, [tf, sig, contrib, just], tech_widths, tech_aligns, row_bg_color=row_color)
            
        # --- PAGE 5: SENTIMENT & RISKS ---
        pdf.add_page()
        
        # Sentiment & News Outlook Section
        pdf.set_font("helvetica", "B", 13)
        pdf.set_text_color(31, 58, 86)
        pdf.cell(0, 8, "Sentiment & News Outlook Analysis", new_x="LMARGIN", new_y="NEXT")
        
        sent_verdict = clean_pdf_text(get_field(sentiment_data, "sentiment_verdict") or "N/A")
        sent_score = clean_pdf_text(get_field(sentiment_data, "sentiment_score") or 0)
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(90, 90, 90)
        pdf.cell(0, 5, f"Verdict: {sent_verdict}  |  Score: {sent_score}/10", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)
        
        themes = get_field(sentiment_data, "key_themes") or []
        themes_text = "• " + "\n• ".join(themes) if themes else "No major themes identified."
        themes_text = clean_pdf_text(themes_text)
        
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(31, 58, 86)
        pdf.cell(0, 4.5, "Core Narrative Themes:", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("helvetica", "", 8.5)
        pdf.set_text_color(60, 60, 60)
        pdf.multi_cell(0, 4.2, themes_text)
        pdf.ln(3)
        
        sent_outlook = clean_pdf_text(get_field(sentiment_data, "future_outlook") or "N/A")
        draw_callout_box(pdf, "Future Outlook Summary", sent_outlook, r=248, g=248, b=250)
        pdf.ln(2)

        # Enriched News Articles Table
        pdf.set_font("helvetica", "B", 10.5)
        pdf.set_text_color(31, 58, 86)
        pdf.cell(0, 6, "Analyzed News Articles & Events", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1.5)
        
        news_widths = [60, 25, 20, 85]
        news_aligns = ["L", "L", "C", "L"]
        draw_table_row(pdf, ["Title / Headline", "Source", "Sentiment", "Summary"], news_widths, news_aligns, is_header=True)
        
        news_articles = get_field(sentiment_data, "news_articles") or []
        for i, article in enumerate(news_articles):
            title = clean_pdf_text(get_field(article, "title") or "N/A")
            src = clean_pdf_text(get_field(article, "source") or "N/A")
            sent_contrib = clean_pdf_text(get_field(article, "sentiment_contribution") or "N/A")
            summary = clean_pdf_text(get_field(article, "summary") or "N/A")
            
            row_color = (250, 252, 254) if i % 2 == 0 else (255, 255, 255)
            draw_table_row(pdf, [title, src, sent_contrib, summary], news_widths, news_aligns, row_bg_color=row_color)
        pdf.ln(4)
        
        # Risks Section (draw_callout_box or normal list)
        pdf.set_font("helvetica", "B", 13)
        pdf.set_text_color(192, 57, 43) # Red for risks
        pdf.cell(0, 8, "Key Risks & Catalysts (Merger Assessment)", new_x="LMARGIN", new_y="NEXT")
        
        risks = get_field(thesis, "risk_factors") or []
        risks_text = "• " + "\n• ".join(risks) if risks else "No major risk factors identified."
        risks_text = clean_pdf_text(risks_text)
        pdf.set_font("helvetica", "", 8.5)
        pdf.set_text_color(60, 60, 60)
        pdf.multi_cell(0, 4.2, risks_text)
        
        # Save PDF report
        file_path = f"{ticker}_Quant_Report.pdf"
        try:
            pdf_bytes = bytes(pdf.output())
            artifact_part = types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf")
            version = await tool_context.save_artifact(filename=file_path, artifact=artifact_part)
            return f"Successfully saved PDF report to ADK artifact service as '{file_path}' (version {version})"
        except Exception as ae:
            return f"Failed to save PDF report as ADK artifact: {str(ae)}"
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"Error exporting PDF: {str(e)}"

async def export_scoring_to_excel(ticker: str, tool_context: ToolContext) -> str:
    """
    Exports the multi-timeframe scoring matrix and indicators to an Excel spreadsheet.
    
    Args:
        ticker (str): The stock ticker symbol.
    """
    try:
        state = tool_context.session.state
        indicator_data = state.get("indicator_final_strategy")
        fundamental_data = state.get("fundamental_results")
        
        if not indicator_data:
            return "Error: No indicator strategy data found in session state."
            
        wb = openpyxl.Workbook()
        
        # Sheet 1: Technical Indicators & Scoring
        ws1 = wb.active
        ws1.title = "Technical Scoring"
        
        # General Info
        ws1.append(["Ticker", ticker])
        ws1.append([])
        
        ws1.append(["Short-Term Verdict", get_field(indicator_data, "short_term_recommendation") or "N/A"])
        ws1.append(["Long-Term Verdict", get_field(indicator_data, "long_term_recommendation") or "N/A"])
        ws1.append(["Universal Net Score", get_field(indicator_data, "total_score") or 0])
        ws1.append(["Weekly RSI Overextension Triggered", get_field(indicator_data, "overextension_override_applied") or False])
        ws1.append([])
        
        # Timeframe Breakdown Table
        ws1.append(["Timeframe", "Signal", "Score Contribution", "Justification"])
        breakdown = get_field(indicator_data, "timeframe_breakdown") or []
        for item in breakdown:
            ws1.append([
                get_field(item, "timeframe") or "",
                get_field(item, "signal") or "",
                get_field(item, "score_contribution") or 0,
                get_field(item, "justification") or ""
            ])
            
        # Advanced Technical Summaries
        ws1.append([])
        ws1.append(["Williams Alligator Analysis Summary", get_field(indicator_data, "alligator_status") or "N/A"])
        ws1.append(["RSI Divergence Analysis Summary", get_field(indicator_data, "divergence_status") or "N/A"])
        ws1.append(["Macro Context & Market Structure Synthesis", get_field(indicator_data, "macro_summary") or "N/A"])
            
        # Sheet 2: Fundamental Valuation & Historical Metrics
        if fundamental_data:
            ws2 = wb.create_sheet(title="Fundamental Scoring")
            ws2.append(["Fundamental Score (out of 10)", get_field(fundamental_data, "fundamental_score") or "N/A"])
            ws2.append(["Valuation Verdict", get_field(fundamental_data, "valuation_verdict") or "N/A"])
            ws2.append(["Calculated Revenue CAGR (%)", get_field(fundamental_data, "revenue_cagr") or 0.0])
            ws2.append(["EPS Trend Verdict", get_field(fundamental_data, "eps_trend_verdict") or "N/A"])
            ws2.append(["Catalyst Summary", get_field(fundamental_data, "catalyst_summary") or ""])
            ws2.append([])
            
            # Programmatic Score Breakdown Table
            ws2.append(["Fundamental Scoring Component", "Score Contribution"])
            breakdown = get_field(fundamental_data, "score_breakdown") or {}
            ws2.append(["Revenue CAGR growth points", get_field(breakdown, "cagr_points") or 0])
            ws2.append(["EPS trend growth points", get_field(breakdown, "eps_points") or 0])
            ws2.append(["Free Cash Flow health points", get_field(breakdown, "fcf_points") or 0])
            ws2.append(["Valuation PE growth points", get_field(breakdown, "pe_growth_points") or 0])
            ws2.append(["Valuation PEG points", get_field(breakdown, "peg_points") or 0])
            ws2.append(["High PE / Low Growth penalty", get_field(breakdown, "high_pe_penalty") or 0])
            ws2.append(["Unprofitable cash burn penalty", get_field(breakdown, "cash_burn_penalty") or 0])
            ws2.append([])
            
            metrics = get_field(fundamental_data, "valuation_metrics") or {}
            ws2.append(["Valuation Metric", "Value"])
            ws2.append(["Trailing P/E", get_field(metrics, "trailing_pe") or "N/A"])
            ws2.append(["Forward P/E", get_field(metrics, "forward_pe") or "N/A"])
            ws2.append(["PEG Ratio", get_field(metrics, "peg_ratio") or "N/A"])
            ws2.append(["Price to Book", get_field(metrics, "price_to_book") or "N/A"])
            ws2.append([])
            
            ws2.append(["Fiscal Year", "Revenue (M)", "Net Income (M)", "Free Cash Flow (M)", "Basic EPS"])
            historical = get_field(fundamental_data, "timeframe_breakdown") or []
            for item in historical:
                ws2.append([
                    get_field(item, "fiscal_year") or "",
                    get_field(item, "revenue_millions") or 0.0,
                    get_field(item, "net_income_millions") or 0.0,
                    get_field(item, "free_cash_flow_millions") or 0.0,
                    get_field(item, "eps") or 0.0
                ])
                
        # Sheet 3: Sentiment & News Outlook
        sentiment_data = state.get("sentiment_results")
        if sentiment_data:
            ws3 = wb.create_sheet(title="Sentiment Analysis")
            ws3.append(["Sentiment Verdict", get_field(sentiment_data, "sentiment_verdict") or "N/A"])
            ws3.append(["Sentiment Score (out of 10)", get_field(sentiment_data, "sentiment_score") or 0])
            ws3.append(["Future Outlook Summary", get_field(sentiment_data, "future_outlook") or ""])
            ws3.append([])
            
            ws3.append(["Core Themes & Narratives"])
            themes = get_field(sentiment_data, "key_themes") or []
            for theme in themes:
                ws3.append([theme])
            ws3.append([])
                
            ws3.append(["Analyzed Key News Articles"])
            ws3.append(["Title", "Source", "Date/Time", "Sentiment Contribution", "Summary", "URL"])
            news_articles = get_field(sentiment_data, "news_articles") or []
            for article in news_articles:
                ws3.append([
                    get_field(article, "title") or "N/A",
                    get_field(article, "source") or "N/A",
                    get_field(article, "published_at") or "N/A",
                    get_field(article, "sentiment_contribution") or "N/A",
                    get_field(article, "summary") or "N/A",
                    get_field(article, "url") or ""
                ])
                
        file_path = f"{ticker}_Scoring_Breakdown.xlsx"
        try:
            import io
            stream = io.BytesIO()
            wb.save(stream)
            excel_bytes = stream.getvalue()
            artifact_part = types.Part.from_bytes(
                data=excel_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            version = await tool_context.save_artifact(filename=file_path, artifact=artifact_part)
            return f"Successfully saved Excel scoring data to ADK artifact service as '{file_path}' (version {version})"
        except Exception as ae:
            return f"Failed to save Excel as ADK artifact: {str(ae)}"
    except Exception as e:
        return f"Error exporting Excel: {str(e)}"

def generate_report_markdown(state: dict) -> str:
    """Programmatically generates the Markdown quant report from structured state variables."""
    thesis = state.get("master_thesis_json") or {}
    pf = state.get("price_feed_results") or {}
    fund = state.get("fundamental_results") or {}
    fund_val = fund.get("valuation_metrics") or {}
    ind = state.get("indicator_final_strategy") or {}
    sent = state.get("sentiment_results") or {}
    
    # Format fundamental timeframe breakdown rows
    fund_tf_lines = []
    for tf in fund.get("timeframe_breakdown", []):
        yr = tf.get("fiscal_year", "N/A")
        rev = tf.get("revenue_millions", "N/A")
        ni = tf.get("net_income_millions", "N/A")
        fcf = tf.get("free_cash_flow_millions", "N/A")
        eps = tf.get("eps", "N/A")
        fund_tf_lines.append(f"| **{yr}** | ${rev}M | ${ni}M | ${fcf}M | ${eps} |")
    fund_tf_str = "\n".join(fund_tf_lines)
    
    # Format indicator timeframe breakdown rows
    ind_tf_lines = []
    for tf in ind.get("timeframe_breakdown", []):
        tfx = tf.get("timeframe", "N/A")
        sig = tf.get("signal", "N/A")
        contrib = tf.get("score_contribution", 0)
        just = tf.get("justification", "N/A")
        ind_tf_lines.append(f"| **{tfx}** | **{sig}** | **{contrib}** | {just} |")
    ind_tf_str = "\n".join(ind_tf_lines)
    
    # Format technical sub-indicator scoring breakdown
    calc_scores = state.get("indicator_calculated_scores") or {}
    breakdown_st = calc_scores.get("breakdown", {}).get("short_term", {})
    breakdown_lt = calc_scores.get("breakdown", {}).get("long_term", {})
    tech_breakdown_lines = []
    for k, v in breakdown_st.items():
        name = f"Short-Term {k.replace('_', ' ').title()}"
        tech_breakdown_lines.append(f"| {name} | **{v}** |")
    for k, v in breakdown_lt.items():
        name = f"Long-Term {k.replace('_', ' ').title()}"
        tech_breakdown_lines.append(f"| {name} | **{v}** |")
    tech_breakdown_str = "\n".join(tech_breakdown_lines) if tech_breakdown_lines else "| N/A | N/A |"
    
    # Format fundamental score breakdown
    fund_breakdown = fund.get("score_breakdown", {})
    fund_breakdown_lines = []
    if fund_breakdown:
        for k, v in fund_breakdown.items():
            name = k.replace('_', ' ').title()
            fund_breakdown_lines.append(f"| {name} | **{v}** |")
    fund_breakdown_str = "\n".join(fund_breakdown_lines) if fund_breakdown_lines else "| N/A | N/A |"

    # Format narrative themes
    sent_themes_lines = []
    for theme in sent.get("key_themes", []):
        sent_themes_lines.append(f"* {theme}")
    sent_themes_str = "\n".join(sent_themes_lines) if sent_themes_lines else "* N/A"
    
    # Format analyzed news articles table
    news_lines = []
    news_articles = sent.get("news_articles", [])
    for article in news_articles:
        title = article.get("title", "N/A")
        source = article.get("source", "N/A")
        pub_at = article.get("published_at", "N/A")
        sent_contrib = article.get("sentiment_contribution", "N/A")
        summary = article.get("summary", "N/A")
        url = article.get("url")
        url_str = f" ([Link]({url}))" if url else ""
        news_lines.append(f"| {title} | {source} | {pub_at} | **{sent_contrib}** | {summary}{url_str} |")
    news_table_str = "\n".join(news_lines) if news_lines else "| N/A | N/A | N/A | N/A | N/A |"

    # Format risks
    risks_lines = []
    for risk in thesis.get("risk_factors", []):
        risks_lines.append(f"* {risk}")
    risks_str = "\n".join(risks_lines) if risks_lines else "* N/A"
    
    return f"""# 🏛️ Institutional Quant Report: {thesis.get('ticker', 'N/A')}
## 🎯 Overall Thesis: **{thesis.get('final_verdict', 'N/A')}** | Conviction: **{thesis.get('conviction_level', 'N/A')}**

---

### 💸 Live Market Price Feed
* **Current Price:** ${pf.get('current_price', 'N/A')}
* **Previous Close:** ${pf.get('previous_close', 'N/A')}
* **Daily Range:** {pf.get('daily_range', 'N/A')}
* **52-Week Range:** {pf.get('fifty_two_week_range', 'N/A')}
* **Daily Volume:** {pf.get('volume', 'N/A')}

---

### 🧠 Strategist Synthesis
> {thesis.get('synthesis_summary', 'N/A')}

---

### 📊 Fundamental Health & Valuation
**Health Score:** {fund.get('fundamental_score', 'N/A')} (Scale: -10 to +10) | **Valuation:** {fund.get('valuation_verdict', 'N/A')}

**Catalyst Summary:**
> {fund.get('catalyst_summary', 'N/A')}

**Programmatic Fundamental Scoring Breakdown:**
| Fundamental Component | Score Contribution |
| :--- | :---: |
{fund_breakdown_str}

| Valuation Metric | Value |
| :--- | :--- |
| **Trailing P/E** | {fund_val.get('trailing_pe', 'N/A')} |
| **Forward P/E** | {fund_val.get('forward_pe', 'N/A')} |
| **PEG Ratio** | {fund_val.get('peg_ratio', 'N/A')} |
| **Price-to-Book** | {fund_val.get('price_to_book', 'N/A')} |

**Historical Performance:**
* **Revenue CAGR:** {fund.get('revenue_cagr', 'N/A')}%
* **Core EPS Trend:** {fund.get('eps_trend_verdict', 'N/A')}

| Fiscal Year | Revenue (M) | Net Income (M) | Free Cash Flow (M) | Basic EPS |
| :---: | :---: | :---: | :---: | :---: |
{fund_tf_str}

---

### 📈 Technical Topology & Momentum
**Short-Term Verdict:** {ind.get('short_term_recommendation', 'N/A')}
**Long-Term Verdict:** {ind.get('long_term_recommendation', 'N/A')}
**Total Score:** {ind.get('total_score', 'N/A')}

**Williams Alligator Analysis:**
> {ind.get('alligator_status', 'N/A')}

**RSI Divergence Analysis:**
> {ind.get('divergence_status', 'N/A')}

**Macro Context:**
> {ind.get('macro_summary', 'N/A')}
> *Overextension Override Triggered: {ind.get('overextension_override_applied', 'N/A')}*

**Programmatic Technical Scoring Breakdown:**
| Technical Component | Score Contribution |
| :--- | :---: |
{tech_breakdown_str}

**Multi-Timeframe Breakdown:**
| Timeframe | Signal | Score Contribution | Technical Justification |
| :--- | :---: | :---: | :--- |
{ind_tf_str}

---

### 📰 Sentiment & News Outlook
* **Sentiment Verdict:** **{sent.get('sentiment_verdict', 'N/A')}** | Score: **{sent.get('sentiment_score', 'N/A')}/10**
* **Core Narrative Themes:**
{sent_themes_str}
* **Future Outlook & Catalyst Summary:**
> {sent.get('future_outlook', 'N/A')}

**Analyzed Key News Articles:**
| Title | Source | Date | Sentiment | Summary |
| :--- | :--- | :--- | :---: | :--- |
{news_table_str}

---

### ⚠️ Risk Factors & Catalysts (Merger Assessment)
**Key Risks to Thesis:**
{risks_str}"""


